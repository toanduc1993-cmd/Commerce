"""
import_pr_mto_from_packages.py

Bulk import PR items từ folder '10 TH-MUA SẮM CÁC GÓI' (IBSHI vault) vào DB platform.

Source: 29 file PR Excel format MTO (Material Take Off) — sheet 'MTO'
Target: PrDetail table với statusFlag='Chờ báo giá'

MTO format (chuẩn IBSHI):
  - Row 4-6: header info (Project Code, Ref. No, Date)
  - Row 8-10: column headers (3-row merged)
  - Row 11+: data items, mỗi section A/B/C/D/E là 1 nhóm vật tư
  - Section letter → materialGroupCode: A=VTC, B=VPK, C=VDK, D=VBP, E=VTH

Cách dùng:
  python3 import_pr_mto_from_packages.py --dry-run     # preview
  python3 import_pr_mto_from_packages.py --apply       # commit
  python3 import_pr_mto_from_packages.py --apply --limit-files 3   # giới hạn N file

Plan:
  1. Walk folder, tìm file PR mới nhất theo (gói, REV)
  2. Parse từng file → list items
  3. Cho mỗi file: tạo Project (nếu chưa có) + 1 prId UUID + N PrDetail rows
  4. statusFlag default = 'Chờ báo giá'
  5. Audit log
"""
import os
import sys
import re
import uuid
import argparse
import json
import unicodedata
from datetime import datetime
from pathlib import Path
from collections import defaultdict


def nfc(s):
    """macOS APFS dùng NFD; normalize sang NFC để regex match được tiếng Việt."""
    return unicodedata.normalize("NFC", str(s)) if s else s

# Lazy import — chỉ load khi cần
try:
    import openpyxl
except ImportError:
    print("Cần `pip3 install openpyxl psycopg2-binary` trước.", file=sys.stderr)
    sys.exit(1)

try:
    import psycopg2
    import psycopg2.extras
except ImportError:
    print("Cần `pip3 install psycopg2-binary` trước.", file=sys.stderr)
    sys.exit(1)


# ─── Config ──────────────────────────────────────────────────────────────────
VAULT_ROOT = Path(
    "/Users/trinhhuuhung/Desktop/HUNGAI/HUNGTH OBSIDIAN V/HUNGTH OBSIDIAN/IBSHI/mua-hang/00.DATA/10 TH-MUA SẮM CÁC GÓI"
)
DB_CONN = "host=127.0.0.1 port=54321 user=vpi_user password=VpiProcurement2026! dbname=vpi_procurement"

# Map section letter → materialGroupCode (theo bidcode v2 catalog)
SECTION_TO_MAT = {
    "A": "VTC",  # Main material → Thép chính
    "B": "VPK",  # Sub material → Phụ kiện
    "C": "VDK",  # Đóng kiện
    "D": "VBP",  # Biện pháp
    "E": "VTH",  # Tiêu hao
    "F": "VTS",  # Sơn
    "G": "VTP",  # Dự phòng
}


# ─── File discovery ─────────────────────────────────────────────────────────
def find_latest_pr_per_package():
    """Tìm file PR mới nhất cho mỗi gói theo (latest 'Cập nhật' folder, highest REV).
    Returns list of (package_id, file_path)."""
    pr_files = []
    for fp in VAULT_ROOT.rglob("*(PR)*.xlsx"):
        if "~$" in fp.name:
            continue  # skip Excel temp lock
        path_str = nfc(fp)
        name_nfc = nfc(fp.name)
        # Extract package ID from path "Gói thầu XXX"
        m = re.search(r"Gói thầu (\S+?)/", path_str)
        if not m:
            continue
        pkg = m.group(1).strip()
        # Skip "phụ kiện", "RUBBER" — đó là PR sub-material/special, để file MTO main thôi
        if "Vật tư phụ" in name_nfc or "Phụ kiện" in path_str or "RUBBER" in name_nfc.upper():
            continue
        # Extract REV number from filename: "...REV 04..." or "...Rev.01..."
        rev_match = re.search(r"REV\s*\.?\s*(\d+)", name_nfc, re.IGNORECASE)
        rev = int(rev_match.group(1)) if rev_match else 0
        # Cập nhật date for ordering
        cn_match = re.search(r"Cập nhật (\d+)-(\d+)-(\d+)", path_str)
        if cn_match:
            d, m_, y = cn_match.groups()
            cap_date = int(f"{y:>04}{m_:>02}{d:>02}")
        else:
            cap_date = 0
        pr_files.append((pkg, rev, cap_date, fp))

    # Skip duplicate "-CLOSE" packages (same content as parent)
    pr_files = [t for t in pr_files if not t[0].endswith("-CLOSE")]

    # Group by package, keep highest cap_date + rev
    by_pkg = defaultdict(list)
    for pkg, rev, cap, fp in pr_files:
        by_pkg[pkg].append((cap, rev, fp))

    latest = []
    for pkg, lst in by_pkg.items():
        lst.sort(key=lambda x: (x[0], x[1]), reverse=True)
        latest.append((pkg, lst[0][2]))
    latest.sort()
    return latest


# ─── Parse PR Excel ──────────────────────────────────────────────────────────
def parse_pr_file(fp: Path):
    """Parse 1 PR Excel file format MTO.
    Returns: { project_code, ref_no, items: [...] }"""
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"open fail: {e}"}

    # Sheet name có thể là 'MTO' hoặc 'PR' hoặc khác — pick first sheet with >50 rows
    ws = None
    for sn in wb.sheetnames:
        candidate = wb[sn]
        if candidate.max_row > 30:
            ws = candidate
            break
    if ws is None:
        ws = wb[wb.sheetnames[0]]

    # Read first 12 rows to extract header info
    head_rows = list(ws.iter_rows(max_row=12, values_only=True))
    head_text = " ".join(str(c) for r in head_rows for c in r if c)

    # Extract project_code (search "Project Code/ Mã dự án:" → value follows)
    proj_match = re.search(r"Mã\s*dự\s*án[:\s]+([^\s\n]+)", head_text)
    project_code = proj_match.group(1).strip() if proj_match else None

    # Extract Ref. No / Mã tài liệu
    ref_match = re.search(r"(?:Ref\.?\s*No|Mã\s*tài\s*liệu)[:\s/]+([A-Z0-9\-\.\s]+REV[\s.]*\d+)", head_text, re.IGNORECASE)
    ref_no = ref_match.group(1).strip() if ref_match else fp.stem

    # Find header row index (row có 'STT' và 'Profile')
    header_row_idx = None
    for i, r in enumerate(head_rows):
        cells = [str(c).lower() if c else "" for c in r]
        joined = " ".join(cells)
        if ("stt" in joined or "item" in joined) and ("profile" in joined or "vật tư" in joined):
            header_row_idx = i
            break
    if header_row_idx is None:
        header_row_idx = 8  # fallback

    # Detect schema variant: tìm vị trí cột 'Grade' trong header row
    #   Variant A (no dimension): Grade at col 3, Unit at 4, UW at 5
    #   Variant B (with DIMENSION): Grade at col 6, Unit at 7, UW at 8
    header_row = head_rows[header_row_idx] if header_row_idx < len(head_rows) else []
    grade_col = None
    for ci, cv in enumerate(header_row):
        if cv and "grade" in str(cv).lower():
            grade_col = ci
            break
    if grade_col is None:
        grade_col = 6  # default to variant B
    has_dimension = grade_col >= 5

    # Schema map: offsets cho từng cột data
    schema = {
        "grade": grade_col,
        "uom": grade_col + 1,
        "unitWeight": grade_col + 2,
        "netQty": grade_col + 3,
        "netWeight": grade_col + 4,
        # Cột Current Ordered nằm cuối:
        # Variant B (068): cols 13-14
        # Variant A (009/075): cols 10-11
        # Variant A2 (063): cols 12-13 (extra Total Ordered)
        "reqQty": grade_col + 7,  # fallback; sẽ try multiple
        "reqWeight": grade_col + 8,
    }

    # Parse items
    # openpyxl rows: 1-indexed. header_row_idx is 0-indexed (from enumerate of head_rows).
    # MTO format: header row at openpyxl (header_row_idx+1), sub-header (+1), col-num row (+1),
    # then section "A." row (+1), then real items.
    # min_row = header_row_idx + 4 (1-indexed) starts at section header row.
    items = []
    current_section = "A"
    for row in ws.iter_rows(min_row=header_row_idx + 4, values_only=True):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        stt = str(row[0]).strip() if row[0] else ""

        # Section header row: chỉ có letter "A." hoặc "B." ở col 0
        section_match = re.match(r"^([A-G])\.?\s*$", stt) if stt else None
        if section_match:
            current_section = section_match.group(1)
            continue

        # Item code patterns:
        #   "A-1", "A.1", "A1", "A48" — letter + optional separator + digit (+ optional letter)
        #   "1", "2" — pure digit (use current_section context)
        m_letter = re.match(r"^([A-G])\s*[-\.]?\s*(\d+[A-Z]?)$", stt) if stt else None
        m_digit = re.match(r"^(\d+)$", stt) if stt else None
        if m_letter:
            section = m_letter.group(1)
            item_seq = m_letter.group(2)
            current_section = section
        elif m_digit:
            section = current_section
            item_seq = m_digit.group(1)
        else:
            continue  # subtotal/footer/empty

        def cell(i):
            return row[i] if i < len(row) else None

        def num(v):
            try:
                if v is None or v == "":
                    return None
                return float(v)
            except (TypeError, ValueError):
                return None

        def text(v, maxlen=200):
            if v is None:
                return None
            s = str(v).strip()
            return s[:maxlen] if s else None

        item_name = text(cell(1))
        profile = text(cell(2))
        # Variant B: append dimension cols (3-5) to profile
        if has_dimension:
            dim_parts = []
            for di in (3, 4, 5):
                dv = cell(di)
                if dv is not None and str(dv).strip() not in ("", "0"):
                    dim_parts.append(str(dv).strip())
            if dim_parts and profile:
                profile = f"{profile} ({' x '.join(dim_parts)})"

        # Try multiple positions cho reqQty/reqWeight (Current Ordered cột cuối)
        # Logic: scan từ phải sang trái, tìm cặp (Qty, Weight) > 0
        # Fallback: dùng netQty/netWeight
        req_qty = None
        req_weight = None
        # Variant B (068/071): cols 13-14
        # Variant A (009/075): cols 10-11
        # Variant A2 (063): cols 12-13
        # Strategy: lấy 2 cột phải nhất có giá trị
        for pair in [(13, 14), (12, 13), (10, 11), (11, 12), (8, 9)]:
            q = num(cell(pair[0]))
            w = num(cell(pair[1]))
            if q is not None and q > 0:
                req_qty = q
                req_weight = w
                break

        if req_qty is None:
            req_qty = num(cell(schema["netQty"])) or 1.0
            req_weight = num(cell(schema["netWeight"]))

        items.append({
            "section": section,
            "matGroupCode": SECTION_TO_MAT.get(section, "ALL"),
            "subGroup": f"{SECTION_TO_MAT.get(section, 'ALL')}01",
            "itemCode": f"{ref_no.split('-')[1] if '-' in ref_no else 'PKG'}-{section}-{item_seq}",
            "itemName": item_name or f"Item {stt}",
            "profile": profile,
            "grade": text(cell(schema["grade"])),
            "uom": text(cell(schema["uom"]), 16),
            "unitWeight": num(cell(schema["unitWeight"])),
            "netQty": num(cell(schema["netQty"])),
            "netWeight": num(cell(schema["netWeight"])),
            "reqQty": req_qty,
            "reqWeight": req_weight,
        })

    return {
        "project_code": project_code,
        "ref_no": ref_no,
        "items": items,
        "sheet_name": ws.title,
    }


# ─── DB import ───────────────────────────────────────────────────────────────
def get_or_create_project(cur, pkg_id, ref_no):
    """Map package ID → Project. Tạo Project nếu chưa có."""
    # Try existing project codes that contain the package number
    # Package "068" → look for project code containing "I-068" or "068"
    cur.execute(
        """SELECT id, code, name FROM "Project"
           WHERE code ILIKE %s OR code ILIKE %s OR code ILIKE %s
           LIMIT 1""",
        (f"%-{pkg_id}", f"%I-{pkg_id}%", f"%{pkg_id}-%"),
    )
    row = cur.fetchone()
    if row:
        return row[0], row[1]

    # Tạo project mới
    new_id = str(uuid.uuid4())
    code = f"PKG-{pkg_id}"
    name = f"Gói thầu {pkg_id}" + (f" (Ref {ref_no})" if ref_no else "")
    cur.execute(
        """INSERT INTO "Project" (id, code, name, "createdAt", "updatedAt")
           VALUES (%s, %s, %s, NOW(), NOW())""",
        (new_id, code, name),
    )
    return new_id, code


def import_file(cur, pkg_id, parsed, apply, source_file):
    """Insert 1 PR file vào DB. Returns dict stats."""
    if not parsed.get("items"):
        return {"items": 0, "skipped": "no items parsed"}

    project_id, project_code = get_or_create_project(cur, pkg_id, parsed.get("ref_no"))

    # Tạo PurchaseRequisition row (FK target của PrDetail.prId)
    pr_id = str(uuid.uuid4())
    pr_ref = (parsed.get("ref_no") or f"PKG-{pkg_id}")[:100]
    cur.execute('SELECT id FROM "PurchaseRequisition" WHERE "prRef"=%s LIMIT 1', (pr_ref,))
    if cur.fetchone():
        pr_ref = f"{pr_ref}-{datetime.now().strftime('%Y%m%d%H%M%S')}"[:100]
    cur.execute(
        """INSERT INTO "PurchaseRequisition" (id, "projectId", "prRef", department, client, status, "createdAt", "updatedAt")
           VALUES (%s, %s, %s, %s, %s, %s, NOW(), NOW())""",
        (
            pr_id,
            project_id,
            pr_ref,
            "ENGINEERING",
            f"Imported from {os.path.basename(source_file)}"[:200],
            "DRAFT",
        ),
    )

    items = parsed["items"]
    inserted = 0
    for i, it in enumerate(items):
        pr_detail_id = str(uuid.uuid4())
        cur.execute(
            """INSERT INTO "PrDetail" (
                id, "prId", "materialGroupCode", "materialSubGroupCode",
                "itemCode", "itemName", profile, grade, uom,
                "unitWeight", "netQty", "netWeight", "reqQty", "reqWeight",
                "remainQty", "remainWeight", "toBuyQty", "toBuyWeight",
                urgency, "statusFlag", remarks,
                "createdAt", "updatedAt"
            ) VALUES (
                %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s,
                %s, %s, %s,
                NOW(), NOW()
            )""",
            (
                pr_detail_id,
                pr_id,
                it["matGroupCode"],
                it.get("subGroup"),
                it["itemCode"][:50],
                (it.get("itemName") or "")[:200],
                (it.get("profile") or "")[:200],
                (it.get("grade") or "")[:100],
                (it.get("uom") or "pcs")[:16],
                it.get("unitWeight") or 0,
                it.get("netQty") or 0,
                it.get("netWeight") or 0,
                it.get("reqQty") or 1.0,
                it.get("reqWeight") or 0,
                it.get("reqQty") or 1.0,  # remainQty = reqQty initially
                it.get("reqWeight") or 0,
                it.get("reqQty") or 1.0,  # toBuyQty
                it.get("reqWeight") or 0,
                "Normal",
                "Chờ báo giá",
                f"Imported from {os.path.basename(source_file)}"[:500],
            ),
        )
        inserted += 1

    return {
        "items": inserted,
        "pr_id": pr_id,
        "project_id": project_id,
        "project_code": project_code,
    }


# ─── Main ────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--apply", action="store_true")
    ap.add_argument("--limit-files", type=int, default=0, help="0 = all")
    ap.add_argument("--filter-pkg", type=str, default="", help="Chỉ import packages khớp pattern")
    args = ap.parse_args()
    if not args.dry_run and not args.apply:
        ap.error("Cần --dry-run hoặc --apply")

    print(f"[{datetime.now().isoformat()}] Import PR MTO từ '10 TH-MUA SẮM CÁC GÓI'")
    print(f"  Mode: {'DRY RUN' if args.dry_run else 'APPLY'}")
    print(f"  Vault: {VAULT_ROOT}")

    if not VAULT_ROOT.exists():
        print(f"  ❌ Vault folder không tồn tại: {VAULT_ROOT}", file=sys.stderr)
        sys.exit(1)

    latest = find_latest_pr_per_package()
    if args.filter_pkg:
        latest = [(p, f) for p, f in latest if args.filter_pkg in p]
    if args.limit_files > 0:
        latest = latest[: args.limit_files]

    print(f"\n  Sẽ xử lý {len(latest)} file PR (1 file mới nhất / gói):")
    for pkg, fp in latest:
        print(f"    [{pkg}] {fp.name}")

    # Parse each
    print("\n  Parsing...")
    parsed_all = []
    for pkg, fp in latest:
        p = parse_pr_file(fp)
        if "error" in p:
            print(f"    ❌ [{pkg}] {p['error']}")
            continue
        item_count = len(p.get("items", []))
        if item_count == 0:
            print(f"    ⚠️  [{pkg}] 0 items parsed (sheet={p.get('sheet_name')})")
            continue
        # Section distribution
        sec_count = defaultdict(int)
        for it in p["items"]:
            sec_count[it["section"]] += 1
        sec_str = ", ".join(f"{k}={v}" for k, v in sorted(sec_count.items()))
        print(f"    ✓ [{pkg}] {item_count} items (sections: {sec_str}) ref={p.get('ref_no','?')[:40]}")
        parsed_all.append((pkg, fp, p))

    total_items = sum(len(p["items"]) for _, _, p in parsed_all)
    print(f"\n  Tổng: {len(parsed_all)} files, {total_items} items sẵn sàng import")

    if args.dry_run:
        print("\n  DRY RUN — không write DB. Sample 3 items:")
        if parsed_all:
            for it in parsed_all[0][2]["items"][:3]:
                print(f"    {it['itemCode']:20} mat={it['matGroupCode']:4} qty={it['reqQty']:8.2f} {it['itemName'][:40]}")
        print("\n  Rerun với --apply để commit.")
        return

    # Apply
    conn = psycopg2.connect(DB_CONN)
    conn.set_client_encoding("UTF8")
    conn.autocommit = False
    try:
        cur = conn.cursor()
        results = []
        for pkg, fp, parsed in parsed_all:
            r = import_file(cur, pkg, parsed, apply=True, source_file=str(fp))
            r["pkg"] = pkg
            r["file"] = fp.name
            results.append(r)
            print(f"    ✓ {pkg}: {r['items']} items → project {r.get('project_code','?')} prId={r.get('pr_id','?')[:8]}")

        # Audit log
        cur.execute(
            """INSERT INTO "AuditLog" (id, action, "entityType", "entityId", details, "createdAt")
               VALUES (%s, %s, %s, %s, %s, NOW())""",
            (
                str(uuid.uuid4()),
                "BULK_IMPORT_PR_MTO",
                "PrDetail",
                "BULK",
                json.dumps({
                    "source": "10 TH-MUA SẮM CÁC GÓI",
                    "filesProcessed": len(results),
                    "totalItems": sum(r["items"] for r in results),
                    "byPackage": {r["pkg"]: r["items"] for r in results},
                }),
            ),
        )

        conn.commit()
        print(f"\n  ✅ COMMIT done — {len(results)} files, {sum(r['items'] for r in results)} items inserted")
    except Exception as e:
        conn.rollback()
        print(f"\n  ❌ ROLLBACK: {e}")
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
